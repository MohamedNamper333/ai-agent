"""Data analysis tools: CSV/JSON/Excel reading, statistics, SQL, data quality, correlation, visualization, time series.

All public methods are static and return formatted strings (not dicts).
"""

import csv
import json
import os
import re
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd


class DataAnalysis:
    @staticmethod
    def _read_data_file(file_path: str):
        p = Path(file_path)
        if not p.exists() or not p.is_file():
            return None
        suffix = p.suffix.lower()
        try:
            if suffix == ".csv":
                return pd.read_csv(p)
            if suffix == ".json":
                try:
                    text = p.read_text(encoding="utf-8")
                except UnicodeDecodeError:
                    text = p.read_text(encoding="latin-1")
                data = json.loads(text)
                if isinstance(data, list):
                    return pd.DataFrame(data)
                if isinstance(data, dict):
                    return pd.DataFrame([data])
                return None
        except Exception:
            return None
        return None

    @staticmethod
    def _compute_col_stats(values: List[Any]) -> Dict[str, Any]:
        nums: List[float] = []
        for v in values:
            if v is None:
                continue
            try:
                s = str(v).strip()
                if s == "" or s.lower() in ("nan", "none", "null"):
                    continue
                nums.append(float(s))
            except (ValueError, TypeError):
                pass
        if nums and len(nums) >= max(1, len(values) // 2 + 1):
            return {
                "type": "numeric",
                "min": min(nums),
                "max": max(nums),
                "avg": sum(nums) / len(nums),
            }
        non_null = sum(
            1 for v in values
            if v is not None and str(v).strip() != "" and str(v).lower() not in ("nan", "none", "null")
        )
        if not values:
            return {"type": "text"}
        return {"type": "text", "unique": len(set(str(v) for v in values)), "non_null": non_null}

    @staticmethod
    def analyze_csv(file_path: str, max_rows: int = 50) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            df = pd.read_csv(file_path)
        except Exception as e:
            return f"Error reading CSV: {e}"
        if len(df.columns) == 0:
            return f"Error: No columns found in CSV {file_path}"
        col_types: Dict[str, str] = {}
        col_stats: Dict[str, Dict[str, Any]] = {}
        for col in df.columns:
            col_types[col] = str(df[col].dtype)
            col_stats[col] = DataAnalysis._compute_col_stats(
                [str(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else None
                 for v in df[col].tolist()]
            )
        lines = [
            f"CSV Analysis: {file_path}",
            f"Total Rows: {len(df)}",
            f"Total Columns: {len(df.columns)}",
            "",
            "Column Types and Stats:",
        ]
        for col in df.columns:
            lines.append(f"  - {col} ({col_types[col]})")
            stats = col_stats.get(col, {})
            for k, v in stats.items():
                if isinstance(v, float):
                    lines.append(f"      {k}: {v:.4f}")
                else:
                    lines.append(f"      {k}: {v}")
        lines.append("")
        lines.append(f"First {min(max_rows, len(df))} rows:")
        lines.append(df.head(max_rows).to_string(index=False))
        return "\n".join(lines)

    @staticmethod
    def analyze_json(file_path: str) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            return f"Error reading JSON: {e}"
        if isinstance(data, list):
            df = pd.DataFrame(data)
            if df.empty:
                return f"JSON Analysis: {file_path}\nList with 0 records"
            lines = [
                f"JSON Analysis: {file_path}",
                f"Total Records: {len(data)}",
                f"Total Fields: {len(df.columns)}",
                "",
                "Field Types:",
            ]
            for col in df.columns:
                lines.append(f"  - {col} ({df[col].dtype})")
            lines.append("")
            lines.append("First 5 records:")
            lines.append(df.head().to_string(index=False))
            return "\n".join(lines)
        if isinstance(data, dict):
            lines = [
                f"JSON Analysis: {file_path}",
                f"Type: object (dict with {len(data)} keys)",
                "",
                "Top-level Keys:",
            ]
            for k, v in data.items():
                if isinstance(v, (dict, list)):
                    lines.append(f"  - {k}: {type(v).__name__} (len={len(v)})")
                else:
                    lines.append(f"  - {k}: {v}")
            return "\n".join(lines)
        return f"JSON Analysis: {file_path}\nScalar value: {data}"

    @staticmethod
    def analyze_text(file_path: str) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
        except Exception as e:
            return f"Error reading text: {e}"
        lines = text.splitlines()
        words = re.findall(r"\b\w+\b", text)
        sentences = re.split(r"[.!?]+", text)
        sentences = [s.strip() for s in sentences if s.strip()]
        return (
            f"Text Analysis: {file_path}\n"
            f"  Characters: {len(text)}\n"
            f"  Lines: {len(lines)}\n"
            f"  Words: {len(words)}\n"
            f"  Sentences: {len(sentences)}\n"
            f"  Avg word length: {(sum(len(w) for w in words) / len(words)) if words else 0:.2f}\n"
            f"  Avg sentence length (words): {(len(words) / len(sentences)) if sentences else 0:.2f}\n"
        )

    @staticmethod
    def analyze_excel(file_path: str, sheet: Optional[str] = None) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            import openpyxl
        except ImportError:
            return "Error: openpyxl is required for Excel analysis"
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as e:
            return f"Error opening Excel: {e}"
        if sheet:
            if sheet not in wb.sheetnames:
                return f"Error: Sheet '{sheet}' not found. Available: {wb.sheetnames}"
            ws = wb[sheet]
        else:
            ws = wb[wb.sheetnames[0]]
        rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
        if not rows:
            return f"Excel Analysis: {file_path}\nEmpty workbook"
        header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        data_rows = rows[1:]
        col_types: Dict[str, str] = {}
        for i, col in enumerate(header):
            col_vals = [r[i] for r in data_rows if i < len(r)]
            nums = sum(1 for v in col_vals if isinstance(v, (int, float)) and not isinstance(v, bool))
            col_types[col] = "numeric" if nums and nums >= len(col_vals) / 2 else "text"
        lines = [
            f"Excel Analysis: {file_path}",
            f"Sheet: {sheet or wb.sheetnames[0]}",
            f"Rows: {len(data_rows)}",
            f"Columns ({len(header)}): {', '.join(header)}",
            "",
            "Column Types:",
        ]
        for col in header:
            lines.append(f"  - {col} ({col_types.get(col, 'unknown')})")
        lines.append("")
        lines.append("First 5 rows:")
        for r in data_rows[:5]:
            lines.append(f"  {r}")
        return "\n".join(lines)

    @staticmethod
    def stats_summary(data_json: str) -> str:
        try:
            data = json.loads(data_json)
        except (json.JSONDecodeError, TypeError):
            return "Error: Provide data as a JSON string (array of numbers)"
        if not isinstance(data, list) or not data:
            return "Error: Data must be a non-empty array of numbers"
        nums: List[float] = []
        for item in data:
            try:
                nums.append(float(item))
            except (ValueError, TypeError):
                continue
        if not nums:
            return "Error: No numeric values found in data"
        sorted_nums = sorted(nums)
        n = len(sorted_nums)
        mean = sum(sorted_nums) / n
        variance = sum((x - mean) ** 2 for x in sorted_nums) / n
        std_dev: Any = "N/A" if n < 5 else f"{variance ** 0.5:.4f}"
        if n % 2 == 0:
            median = (sorted_nums[n // 2 - 1] + sorted_nums[n // 2]) / 2
        else:
            median = sorted_nums[n // 2]
        q1 = sorted_nums[max(0, n // 4)]
        q3 = sorted_nums[min(n - 1, 3 * n // 4)]
        return (
            f"Statistical Summary (n={n}):\n"
            f"  Min: {min(nums):.4f}\n"
            f"  Max: {max(nums):.4f}\n"
            f"  Mean: {mean:.4f}\n"
            f"  Median: {median:.4f}\n"
            f"  Std Dev: {std_dev}\n"
            f"  Q1: {q1:.4f}\n"
            f"  Q3: {q3:.4f}\n"
            f"  Range: {max(nums) - min(nums):.4f}\n"
            f"  Sum: {sum(nums):.4f}\n"
        )

    @staticmethod
    def sql_query(file_path: str, query: str) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            conn = sqlite3.connect(":memory:")
            if file_path.lower().endswith(".csv"):
                with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                if not rows:
                    return "Error: Empty CSV file"
                cols = list(rows[0].keys())
                col_defs = ", ".join(f'"{c}"' for c in cols)
                conn.execute(f"CREATE TABLE data ({col_defs})")
                placeholders = ", ".join("?" for _ in cols)
                for row in rows:
                    conn.execute(
                        f"INSERT INTO data ({col_defs}) VALUES ({placeholders})",
                        [row.get(c, "") for c in cols],
                    )
            elif file_path.lower().endswith(".json"):
                with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    data = [data]
                if not isinstance(data, list) or not data:
                    return "Error: JSON must be an array of objects"
                cols = list(data[0].keys())
                col_defs = ", ".join(f'"{c}"' for c in cols)
                conn.execute(f"CREATE TABLE data ({col_defs})")
                placeholders = ", ".join("?" for _ in cols)
                for row in data:
                    conn.execute(
                        f"INSERT INTO data ({col_defs}) VALUES ({placeholders})",
                        [row.get(c, "") for c in cols],
                    )
            else:
                return f"Error: Unsupported file format for SQL: {file_path}"
            cur = conn.execute(query)
            results = cur.fetchall()
            headers = [d[0] for d in cur.description] if cur.description else []
            lines = [
                f"SQL Result ({len(results)} rows):",
            ]
            if headers:
                lines.append("  " + " | ".join(str(h) for h in headers))
                lines.append("  " + "-" * 40)
            for r in results[:50]:
                lines.append("  " + " | ".join(str(v) for v in r))
            if len(results) > 50:
                lines.append(f"  ... ({len(results) - 50} more rows)")
            return "\n".join(lines)
        except Exception as e:
            return f"Error executing query: {e}"

    @staticmethod
    def analyze_data_quality(file_path: str) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            if file_path.lower().endswith(".csv"):
                df = pd.read_csv(file_path)
            elif file_path.lower().endswith(".json"):
                with open(file_path, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, dict):
                    raw = [raw]
                df = pd.DataFrame(raw)
            else:
                return f"Error: Unsupported format for data quality: {file_path}"
        except Exception as e:
            return f"Error reading file: {e}"
        rows, cols = df.shape
        missing = int(df.isnull().sum().sum())
        duplicates = int(df.duplicated().sum())
        dtypes = {c: str(t) for c, t in df.dtypes.items()}
        lines = [
            f"Data Quality Report: {file_path}",
            f"  Rows: {rows}",
            f"  Columns: {cols}",
            f"  Missing Values: {missing}",
            f"  Duplicates: {duplicates}",
            f"  Data Types: {dtypes}",
        ]
        per_col_missing = {c: int(df[c].isnull().sum()) for c in df.columns}
        per_col_missing = {c: v for c, v in per_col_missing.items() if v > 0}
        if per_col_missing:
            lines.append(f"  Per-column missing: {per_col_missing}")
        return "\n".join(lines)

    @staticmethod
    def correlation_analysis(file_path: str, columns: Optional[str] = None) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            if file_path.lower().endswith(".csv"):
                df = pd.read_csv(file_path)
            elif file_path.lower().endswith(".json"):
                with open(file_path, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, dict):
                    raw = [raw]
                df = pd.DataFrame(raw)
            else:
                return f"Error: Unsupported format for correlation: {file_path}"
        except Exception as e:
            return f"Error reading file: {e}"
        numeric_df = df.select_dtypes(include="number")
        if columns:
            requested = [c.strip() for c in columns.split(",") if c.strip()]
            present = [c for c in requested if c in numeric_df.columns]
            if not present:
                return (
                    f"Error: None of the specified columns found in numeric data: {requested}. "
                    f"Available numeric columns: {list(numeric_df.columns)}"
                )
            numeric_df = numeric_df[present]
        if numeric_df.shape[1] == 0:
            return "No numeric columns to analyze"
        if numeric_df.shape[1] == 1:
            col = numeric_df.columns[0]
            return f"Correlation Analysis: only one numeric column ({col}) — no correlations possible."
        try:
            corr = numeric_df.corr()
        except Exception as e:
            return f"Error in correlation analysis: {e}"
        lines = [
            "Correlation Analysis:",
            f"  Source: {file_path}",
            f"  Numeric Columns: {list(corr.columns)}",
            "",
            "Pearson correlation matrix:",
        ]
        lines.append(corr.to_string())
        return "\n".join(lines)

    @staticmethod
    def generate_visualization(file_path: str, chart_type: str = "bar", column: Optional[str] = None) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            if file_path.lower().endswith(".csv"):
                df = pd.read_csv(file_path)
            elif file_path.lower().endswith(".json"):
                with open(file_path, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, dict):
                    raw = [raw]
                df = pd.DataFrame(raw)
            else:
                return f"Error: Unsupported format for visualization: {file_path}"
        except Exception as e:
            return f"Error reading file: {e}"
        return (
            f"Visualization Specification:\n"
            f"  Source: {file_path}\n"
            f"  Chart Type: {chart_type}\n"
            f"  Column: {column or 'all'}\n"
            f"  Rows: {len(df)}\n"
            f"  Available Columns: {list(df.columns)}\n"
        )

    @staticmethod
    def time_series_analysis(
        file_path: str,
        date_column: Optional[str] = None,
        value_column: Optional[str] = None,
    ) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"
        try:
            if file_path.lower().endswith(".csv"):
                df = pd.read_csv(file_path)
            elif file_path.lower().endswith(".json"):
                with open(file_path, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, dict):
                    raw = [raw]
                df = pd.DataFrame(raw)
            else:
                return f"Error: Unsupported format for time series: {file_path}"
        except Exception as e:
            return f"Error reading file: {e}"
        return (
            f"Time Series Analysis:\n"
            f"  Source: {file_path}\n"
            f"  Rows: {len(df)}\n"
            f"  Date Column: {date_column or 'auto'}\n"
            f"  Value Column: {value_column or 'auto'}\n"
            f"  Available Columns: {list(df.columns)}\n"
        )
